from __future__ import annotations

import shutil
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import mysql.connector
from openpyxl import load_workbook

PROJECT_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_DIR))

from app.config import IMPORT_DIR, PRODUCT_IMAGES_DIR, get_database_config, relative_to_base


# Таблицы очищаются от зависимых к родительским, чтобы импорт можно было повторять.
TABLES_TO_CLEAR = [
    "order_items",
    "orders",
    "products",
    "users",
    "pickup_points",
    "order_statuses",
    "roles",
    "categories",
    "manufacturers",
    "suppliers",
    "units",
]


# SQL-скрипт схемы разбивается на отдельные команды для последовательного выполнения.
def clean_sql(script: str) -> list[str]:
    lines = []
    for line in script.splitlines():
        stripped = line.strip()
        if stripped.startswith("--"):
            continue
        lines.append(line)
    return [statement.strip() for statement in "\n".join(lines).split(";") if statement.strip()]


# Создает базу данных и таблицы из sql/schema.sql.
def run_schema() -> None:
    config = get_database_config(include_database=False)
    sql_path = PROJECT_DIR / "sql" / "schema.sql"
    connection = mysql.connector.connect(**config)
    try:
        cursor = connection.cursor()
        for statement in clean_sql(sql_path.read_text(encoding="utf-8")):
            cursor.execute(statement)
        connection.commit()
    finally:
        connection.close()


# Полностью очищает учебную БД перед повторной загрузкой исходных данных.
def reset_tables(cursor) -> None:
    cursor.execute("SET FOREIGN_KEY_CHECKS = 0")
    for table in TABLES_TO_CLEAR:
        cursor.execute(f"TRUNCATE TABLE {table}")
    cursor.execute("SET FOREIGN_KEY_CHECKS = 1")


# Читает непустые строки Excel-листа без привязки к имени файла.
def workbook_rows(path: Path) -> tuple[int, list[list[Any]]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    rows: list[list[Any]] = []
    for row in sheet.iter_rows(values_only=True):
        values = list(row)
        while values and values[-1] is None:
            values.pop()
        if any(value is not None for value in values):
            rows.append(values)
    return sheet.max_column or 0, rows


# Определяет тип Excel-файла по количеству колонок: товары, пользователи, заказы, пункты.
def find_workbooks() -> dict[str, list[list[Any]]]:
    result: dict[str, list[list[Any]]] = {}
    for path in IMPORT_DIR.glob("*.xlsx"):
        max_column, rows = workbook_rows(path)
        if max_column == 11:
            result["products"] = rows
        elif max_column == 4:
            result["users"] = rows
        elif max_column == 12:
            result["orders"] = rows
        elif max_column == 1:
            result["pickup_points"] = rows

    missing = {"products", "users", "orders", "pickup_points"} - result.keys()
    if missing:
        raise RuntimeError(f"Не найдены исходные книги: {', '.join(sorted(missing))}")
    return result


# Создает запись справочника или возвращает id уже существующей записи.
def get_or_create(cursor, table: str, name: str) -> int:
    cursor.execute(
        f"""
        INSERT INTO {table} (name)
        VALUES (%s)
        ON DUPLICATE KEY UPDATE id = LAST_INSERT_ID(id)
        """,
        (name.strip(),),
    )
    return int(cursor.lastrowid)


# Приводит числовые значения Excel к Decimal для цен и скидок.
def parse_decimal(value: Any) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    return Decimal(str(value).replace(",", "."))


# Невалидные даты из исходников, например 30.02.2024, импортируются как NULL.
def parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


# Исходный заказ хранит позиции одной строкой: артикул, количество, артикул, количество.
def parse_order_items(raw_value: Any) -> list[tuple[str, int]]:
    parts = [part.strip() for part in str(raw_value).split(",") if part.strip()]
    items: list[tuple[str, int]] = []
    for index in range(0, len(parts), 2):
        if index + 1 >= len(parts):
            break
        items.append((parts[index], int(parts[index + 1])))
    return items


# Копирует фото товара в рабочую папку приложения и возвращает относительный путь.
def prepare_photo(photo_name: str | None) -> str:
    if not photo_name:
        return relative_to_base(IMPORT_DIR / "picture.png")

    source = IMPORT_DIR / str(photo_name)
    destination = PRODUCT_IMAGES_DIR / source.name
    if source.exists() and not destination.exists():
        shutil.copy2(source, destination)
    if destination.exists():
        return relative_to_base(destination)
    return relative_to_base(IMPORT_DIR / "picture.png")


# Пункты выдачи в исходном файле идут без отдельного id, поэтому id задается порядком строк.
def import_pickup_points(cursor, rows: list[list[Any]]) -> None:
    for index, row in enumerate(rows, start=1):
        address = str(row[0]).strip()
        cursor.execute(
            "INSERT INTO pickup_points (id, address) VALUES (%s, %s)",
            (index, address),
        )


# Загружает пользователей и возвращает словарь ФИО -> id для привязки заказов к клиентам.
def import_users(cursor, rows: list[list[Any]]) -> dict[str, int]:
    users_by_name: dict[str, int] = {}
    for row in rows[1:]:
        role_id = get_or_create(cursor, "roles", str(row[0]))
        cursor.execute(
            """
            INSERT INTO users (role_id, full_name, login, password)
            VALUES (%s, %s, %s, %s)
            """,
            (role_id, str(row[1]).strip(), str(row[2]).strip(), str(row[3]).strip()),
        )
        users_by_name[str(row[1]).strip()] = int(cursor.lastrowid)
    return users_by_name


# Загружает товары и параллельно наполняет справочники 3НФ.
def import_products(cursor, rows: list[list[Any]]) -> dict[str, int]:
    products_by_article: dict[str, int] = {}
    for row in rows[1:]:
        unit_id = get_or_create(cursor, "units", str(row[2]))
        supplier_id = get_or_create(cursor, "suppliers", str(row[4]))
        manufacturer_id = get_or_create(cursor, "manufacturers", str(row[5]))
        category_id = get_or_create(cursor, "categories", str(row[6]))
        cursor.execute(
            """
            INSERT INTO products (
                article, name, unit_id, price, supplier_id, manufacturer_id,
                category_id, discount_percent, stock_quantity, description, photo_path
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                str(row[0]).strip(),
                str(row[1]).strip(),
                unit_id,
                parse_decimal(row[3]),
                supplier_id,
                manufacturer_id,
                category_id,
                parse_decimal(row[7]),
                int(row[8] or 0),
                str(row[9] or "").strip(),
                prepare_photo(str(row[10]).strip() if len(row) > 10 and row[10] else None),
            ),
        )
        products_by_article[str(row[0]).strip()] = int(cursor.lastrowid)
    return products_by_article


# Создает заказы и отдельные позиции order_items по артикулам товаров.
def import_orders(
    cursor,
    rows: list[list[Any]],
    users_by_name: dict[str, int],
    products_by_article: dict[str, int],
) -> None:
    for row in rows[1:]:
        status_id = get_or_create(cursor, "order_statuses", str(row[7]))
        customer_id = users_by_name.get(str(row[5]).strip())
        cursor.execute(
            """
            INSERT INTO orders (
                id, order_date, delivery_date, pickup_point_id,
                customer_id, receive_code, status_id
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """,
            (
                int(row[0]),
                parse_date(row[2]),
                parse_date(row[3]),
                int(row[4]),
                customer_id,
                int(row[6]) if row[6] is not None else None,
                status_id,
            ),
        )

        order_id = int(row[0])
        for article, quantity in parse_order_items(row[1]):
            product_id = products_by_article.get(article)
            if product_id is None:
                print(f"Пропущена позиция заказа: артикул {article} не найден")
                continue
            cursor.execute(
                """
                INSERT INTO order_items (order_id, product_id, quantity)
                VALUES (%s, %s, %s)
                """,
                (order_id, product_id, quantity),
            )


# Точка входа: создать схему, очистить таблицы и загрузить все исходные файлы.
def main() -> None:
    run_schema()
    workbooks = find_workbooks()
    config = get_database_config(include_database=True)
    connection = mysql.connector.connect(**config)
    try:
        cursor = connection.cursor()
        reset_tables(cursor)
        import_pickup_points(cursor, workbooks["pickup_points"])
        users_by_name = import_users(cursor, workbooks["users"])
        products_by_article = import_products(cursor, workbooks["products"])
        import_orders(cursor, workbooks["orders"], users_by_name, products_by_article)
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()

    print("Готово: база mebelorg создана и данные импортированы.")


if __name__ == "__main__":
    main()
